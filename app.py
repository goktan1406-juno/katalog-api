from flask import Flask, request, jsonify
import base64, zipfile, os, tempfile, json, re
from io import BytesIO
from collections import defaultdict

from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import ImageReader
from PIL import Image
import reportlab

app = Flask(__name__)
# Werkzeug 3.1+ caps non-file form fields (e.g. prev_state_json, which grows as the
# catalog accumulates products) at 500KB by default; raise it so accumulated state
# doesn't start failing once the catalog grows past a handful of products.
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024
app.config['MAX_FORM_MEMORY_SIZE'] = 200 * 1024 * 1024

FONT_MAP = {}
fonts_loaded = False

_REPORTLAB_FONTS = os.path.join(os.path.dirname(reportlab.__file__), 'fonts')
_FONT_CANDIDATES = [
    ('Sans',     os.path.join(_REPORTLAB_FONTS, 'Vera.ttf')),
    ('SansBold', os.path.join(_REPORTLAB_FONTS, 'VeraBd.ttf')),
    ('SansObl',  os.path.join(_REPORTLAB_FONTS, 'VeraIt.ttf')),
    ('Sans',     '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'),
    ('SansBold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
    ('SansObl',  '/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf'),
]

def load_fonts():
    global fonts_loaded, FONT_MAP
    if fonts_loaded: return
    for name, path in _FONT_CANDIDATES:
        if name not in FONT_MAP and os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(name, path))
                FONT_MAP[name] = path
            except: pass
    fonts_loaded = True

def F():  return 'Sans'     if 'Sans'     in FONT_MAP else 'Helvetica'
def FB(): return 'SansBold' if 'SansBold' in FONT_MAP else 'Helvetica-Bold'

W, H  = A4
RED   = colors.HexColor('#E8281A')
DARK  = colors.HexColor('#1C1C1C')
LGRAY = colors.HexColor('#F5F5F5')
MGRAY = colors.HexColor('#CCCCCC')
DGRAY = colors.HexColor('#555555')
WHITE = colors.white

def tw(cv, t, f, s): return cv.stringWidth(t, f, s)

def wrap(cv, text, font, size, max_w):
    words = str(text).split(); lines, line = [], ''
    for w in words:
        t = (line+' '+w).strip()
        if tw(cv,t,font,size) <= max_w: line = t
        else:
            if line: lines.append(line)
            line = w
    if line: lines.append(line)
    return lines

def parse_xlsm(xlsm_bytes, filename=None):
    buf = BytesIO(xlsm_bytes)
    wb  = load_workbook(buf, data_only=True)
    rt  = wb['RANGE TABLE']
    def get(label):
        for row in rt.iter_rows(values_only=True):
            if row[0] and str(row[0]).strip() == label:
                return str(row[1]).strip() if row[1] else ''
        return ''
    def get_raw(label):
        for row in rt.iter_rows(values_only=True):
            if row[0] and str(row[0]).strip() == label:
                v = row[1]
                if isinstance(v, float) and v.is_integer():
                    return str(int(v))
                return str(v).strip() if v else ''
        return ''
    def get_technical_characteristics():
        section_headers = {
            'RANGE TABLE', 'PRODUCT INFORMATION', 'MARKETING INFORMATION', 'MEDIA IDs',
            'TECHNICAL CHARACTERISTICS', 'ENVIRONMENTAL CHARACTERISTICS',
            'WEIGHT AND DIMENSIONS (UNPACKED)', 'REPAIRABILITY', 'LOGISTICS DATA', 'RELATIONS',
        }
        pairs, in_section = {}, False
        for row in rt.iter_rows(values_only=True):
            if not row[0]:
                continue
            k = str(row[0]).strip()
            if k == 'TECHNICAL CHARACTERISTICS':
                in_section = True
                continue
            if k in section_headers:
                in_section = False
                continue
            if in_section and row[1] not in (None, ''):
                v = row[1]
                if isinstance(v, float) and v.is_integer():
                    v = str(int(v))
                pairs[k] = str(v).strip()
        return pairs
    name = translate_name_to_turkish(trim_name_to_core(get('Commercial Name')))
    category = get('PL') or get('Family L1') or 'GENEL'
    if category == 'COOKWARE & BAKEWARE' and filename:
        name = os.path.splitext(os.path.basename(filename))[0]
    product = {
        'ref':        get('Product Reference'),
        'product_id': get_raw('Product Id'),
        'brand':    get('Brand'),
        'name':     name,
        'claim':    get('Key claim'),
        'category': category,
        'series':   get('Family L2') or get('Range name') or get('Range') or get('Series') or ' '.join(name.split()[:2]),
        'benefits': [],
    }
    seen = set()
    for i in ['1 (USP)','2','3','4','5','6','7','8']:
        t = get(f'Benefit title {i}')
        d = get(f'Benefit detail {i}')
        if t and t.lower() not in ('none','') and t not in seen:
            seen.add(t); product['benefits'].append((t,d))
    if product['benefits']:
        product['benefits'] = ensure_benefits_turkish(product['benefits'])
    if not product['benefits']:
        highlights_raw = get('Benefits Highlights') or get('Short description detail')
        if highlights_raw:
            product['benefits'] = summarize_highlights(highlights_raw)
    category_context = f"{product['category']} {get('Family L1')} {product['series']} {product['ref']}"
    tech_bullets = extract_tech_bullets(get_technical_characteristics(), product_name=name,
                                         category_context=category_context)
    if tech_bullets:
        product['benefits'] = product['benefits'][:6 - len(tech_bullets)] + tech_bullets
    product['images_b64'] = extract_images_b64(buf)
    return product

def extract_images_b64(xlsm_buf):
    imgs = []
    xlsm_buf.seek(0)
    try:
        with zipfile.ZipFile(xlsm_buf,'r') as z:
            for path in [f for f in z.namelist() if f.startswith('xl/media/')]:
                data = z.read(path)
                try:
                    im = Image.open(BytesIO(data))
                    if min(im.size) >= 150:
                        if im.mode == 'P': im = im.convert('RGBA')
                        if im.mode == 'RGBA':
                            bg = Image.new('RGB', im.size, (255,255,255))
                            bg.paste(im, mask=im.split()[3]); im = bg
                        else: im = im.convert('RGB')
                        out = BytesIO()
                        im.save(out, 'JPEG', quality=85)
                        imgs.append(base64.b64encode(out.getvalue()).decode())
                        break
                except: pass
    except: pass
    return imgs

def b64_to_reader(b64):
    try:
        return ImageReader(BytesIO(base64.b64decode(b64)))
    except: return None

def draw_page_chrome(cv, page_num, category):
    HDR = 13*mm
    cv.setFillColor(DARK); cv.rect(0, H-HDR, W, HDR, fill=1, stroke=0)
    cv.setFillColor(RED);  cv.rect(0, H-HDR, 3.5*mm, HDR, fill=1, stroke=0)
    cv.setFillColor(WHITE); cv.setFont(FB(), 9)
    cv.drawString(8*mm, H-HDR+4.5*mm, str(category).upper())
    cv.setFont(F(), 7.5); cv.setFillColor(MGRAY)
    cv.drawRightString(W-8*mm, H-HDR+4.5*mm, 'Urun Katalogu 2026')
    FTR = 10*mm
    cv.setFillColor(LGRAY); cv.rect(0, 0, W, FTR, fill=1, stroke=0)
    cv.setStrokeColor(MGRAY); cv.setLineWidth(0.3); cv.line(0, FTR, W, FTR)
    cv.setFillColor(DGRAY); cv.setFont(F(), 6.5)
    cv.drawString(8*mm, 3.5*mm, '2026 TEFAL')
    cv.drawRightString(W-8*mm, 3.5*mm, 'tefal.com.tr')
    cv.setFillColor(DARK); cv.setFont(FB(), 7)
    cv.drawCentredString(W/2, 3.5*mm, f'{page_num} | TEFAL')

def draw_card(cv, x, y, cw, ch, product):
    """3-column card: white image, bold name, tightly packed bullet features."""
    imgs = [b64_to_reader(b) for b in product.get('images_b64', [])]
    imgs = [i for i in imgs if i]

    IMG_H = cw * 0.76

    # Image area
    img_bot = y - IMG_H
    cv.setFillColor(WHITE)
    cv.rect(x, img_bot, cw, IMG_H, fill=1, stroke=0)
    if imgs:
        try:
            cv.drawImage(imgs[0], x, img_bot, cw, IMG_H,
                         preserveAspectRatio=True, anchor='c', mask='auto')
        except: pass

    # Product name — tight gap, compact line spacing
    ty = img_bot - 2.5*mm
    cv.setFillColor(DARK); cv.setFont(FB(), 8)
    name_lines = wrap(cv, product.get('name', ''), FB(), 8, cw)
    for i, ln in enumerate(name_lines[:2]):
        if i == 1 and len(name_lines) > 2:
            ln = ln.rstrip('.,;: ')
        cv.drawString(x, ty, ln)
        ty -= 4*mm

    # Product ID (no extra pre-gap)
    cv.setFillColor(DGRAY); cv.setFont(F(), 6)
    cv.drawString(x, ty, product.get('product_id', ''))
    ty -= 3.5*mm

    # Separator
    cv.setStrokeColor(colors.HexColor('#DEDEDE')); cv.setLineWidth(0.3)
    cv.line(x, ty, x + cw, ty)
    ty -= 2.5*mm

    # Bullet features — tight spacing so they start close to separator
    BUL_X = x + 4*mm
    BUL_W = cw - 4.5*mm
    bottom_limit = y - ch + 3*mm

    for title, _ in product.get('benefits', [])[:6]:
        if ty - 3.5*mm < bottom_limit: break
        text = str(title)
        while tw(cv, text, F(), 6.5) > BUL_W and len(text) > 5:
            text = text[:-2] + '.'
        cv.setFillColor(RED)
        cv.circle(x + 1.8*mm, ty + 0.7*mm, 1.1*mm, fill=1, stroke=0)
        cv.setFillColor(DARK); cv.setFont(F(), 6.5)
        cv.drawString(BUL_X, ty, text)
        ty -= 3.5*mm

def build_pdf(products, output_path, category):
    # Group same-series products together so they appear side by side
    products = sorted(products, key=lambda p: (p.get('series', ''), p.get('name', '')))

    load_fonts()
    cv = canvas.Canvas(output_path, pagesize=A4)
    cv.setTitle(f'TEFAL {category} Katalogu 2026')

    MARGIN   = 12*mm
    HDR_H    = 13*mm
    FTR_H    = 10*mm
    COLS     = 3
    COL_GAP  = 6*mm
    ROW_GAP  = 8*mm

    card_w = (W - 2*MARGIN - (COLS - 1)*COL_GAP) / COLS
    card_h = 95*mm

    usable_h = H - HDR_H - FTR_H - 2*MARGIN
    rows_pp  = max(1, int((usable_h + ROW_GAP) / (card_h + ROW_GAP)))
    per_page = COLS * rows_pp

    product_start_y = H - HDR_H - MARGIN

    def init_page(pnum):
        draw_page_chrome(cv, pnum, category)

    page_num = 1
    init_page(page_num)
    page_i = 0

    for product in products:
        if page_i >= per_page:
            cv.showPage(); page_num += 1
            init_page(page_num)
            page_i = 0

        col = page_i % COLS
        row = page_i // COLS
        x = MARGIN + col * (card_w + COL_GAP)
        y = product_start_y - row * (card_h + ROW_GAP)
        draw_card(cv, x, y, card_w, card_h, product)
        page_i += 1

    cv.showPage()
    cv.save()

# ─── In-memory catalog state ───────────────────────────────────────────────────
CATALOG_STATE = defaultdict(dict)

_ENGLISH_BULLET_HINTS = (
    'the ', ' and ', ' with ', ' for ', ' our ', ' your ', 'cleaning', 'vacuuming',
    'brush', 'effortless', 'deep', 'complete', 'set', 'power', 'technology',
)

def ensure_benefits_turkish(benefits):
    """Translate English-looking bullet titles pulled straight from the raw
    Benefit title/detail fields; Turkish ones pass through unchanged."""
    titles = [t for t, _ in benefits]
    combined = ' '.join(titles).lower()
    if not any(h in combined for h in _ENGLISH_BULLET_HINTS):
        return benefits
    try:
        import anthropic
        client = anthropic.Anthropic()
        prompt = (
            f"Asagidaki {len(titles)} urun ozelligi maddesini incele. Ingilizce "
            "olanlari Turkceye cevir, zaten Turkce olanlari degistirmeden ayni "
            f"birak. Tam olarak {len(titles)} madde don, satir satir, ayni "
            "sirada, numara veya tire koyma, baska hicbir aciklama ekleme.\n\n"
            + '\n'.join(titles)
        )
        message = client.messages.create(
            model="claude-haiku-4-5", max_tokens=400,
            messages=[{"role": "user", "content": prompt}],
        )
        lines = [l.strip(' -•\t') for l in message.content[0].text.strip().split('\n') if l.strip()]
        if len(lines) == len(titles):
            return [(lines[i], benefits[i][1]) for i in range(len(titles))]
        return benefits
    except Exception as e:
        print(f"Benefit translate error: {e}")
        return benefits

def summarize_highlights(highlights_text):
    """Use Claude Haiku to turn a long highlights/description paragraph into short bullet points."""
    try:
        import anthropic
        client = anthropic.Anthropic()
        prompt = (
            "Asagidaki urun tanitim metnini katalog kartinda gosterilecek kisa "
            "maddeler halinde ozetle. Her madde en fazla 6-8 kelime olsun, "
            "toplam 4-6 madde don. Sadece maddeleri satir satir yaz, "
            "numara veya tire koyma, baska hicbir aciklama ekleme.\n\n"
            f"Metin: {highlights_text}"
        )
        message = client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )
        lines = [l.strip(' -•\t') for l in message.content[0].text.strip().split('\n') if l.strip()]
        return [(l, '') for l in lines[:6]]
    except Exception as e:
        print(f"Highlights summarize error: {e}")
        return []


# Category-specific priority hints: (keyword-matcher over category/series/name/ref, guidance
# text, optional bullet count). When a product matches, its guidance is injected into the
# tech-bullet prompt so Claude prioritizes the specs that actually matter for that product
# type over generic ones. Count defaults to 2 when omitted.
CATEGORY_SPEC_HINTS = [
    (('torbal', 'fc'), "ses yuksekligi (dB), toz haznesi kapasitesi, baslik secenekleri, kutudan cikan aksesuarlar"),
    (('dikey', 'sarjli', 'kablosuz supurge'), "air watt gucu, sarjla calisma suresi (dk), aksesuarlar, toz haznesi kapasitesi"),
    (('utu', 'steam iron', 'steam st', 'st station', 'buhar kazanl'),
     "watt gucu, sok buhar (g/dk), surekli buhar cikisi (g/dk), taban tipi/kaplamasi, su haznesi kapasitesi (ml/L)", 4),
    (('tost makine', 'toast'),
     "watt gucu, ekmek kapasitesi, firin fonksiyonu, plaka materyali, cikarilabilir plakalar, "
     "bulasik makinesinde yikanabilirlik, 180 derece izgara pozisyonu, yag toplama haznesi", 4),
    (('handblend', 'stick mixr', 'el blender', 'hand blend'),
     "dograyici hazne kapasitesi, dograyici hazne materyali, watt gucu, "
     "karistirma kabi hazne kapasitesi, bicak sayisi, aksesuarlar", 4),
    (('blender',),
     "watt gucu, hazne kapasitesi, hazne materyali, bicak sayisi, "
     "(sadece fresh express modellerinde) baslik sayisi", 4),
    (('kiyma makine', 'meat mincer', 'mincer'),
     "watt gucu, dakikada kac kg'a kadar kiyma kapasitesi, kac adet paslanmaz celik plaka", 3),
    (('caydanlik', 'tea maker', 'cay makine'),
     "litre cinsinden paslanmaz celik/cam su isitici kapasitesi, litre cinsinden celik/cam demlik kapasitesi, "
     "govde materyali, watt gucu", 4),
    (('kettle', 'su isitici'),
     "watt gucu, litre cinsinden kapasite, kapak ozelligi, "
     "(varsa) kirec filtresi, (varsa) 360 derece donebilir taban", 4),
    (('duzlestir', 'straight'),
     "plaka tipi/materyali (orn. seramik), iyonik/statik azaltici teknoloji, "
     "sicaklik ayari sayisi ve araligi (derece C), 2'si 1 arada kullanim ozelligi "
     "(duzlestirme+bukle), hizli isinma suresi (saniye), asiri isinmayi onleyen koruma", 4),
    (('bathroom scale', 'banyo tart', 'vucut tart'),
     "max kg kapasite, LCD ekran, gr cinsinden hassasiyet, kolay temizlenir cam yuzey, "
     "kilo/yag yuzdesi/vucut kitle indeksi olcumu, kac kisilik hafiza", 4),
    (('tras makine', 'male beauty', 'x series'),
     "wet&dry (su altinda kullanim), titanyum kaplama, RPM motor gucu, "
     "sac/sakal kesim uzunlugu secenekleri (mm), vucut/burun/kulak tras baslik aksesuarlari, "
     "sarjla calisma suresi (dk)", 4),
    (('vantilator', 'fans clas', 'fan clas', 'fans other', 'fan other'),
     "otomatik ve dikey salinim secenekleri, hiz kademesi ve turbo, uzaktan kumanda, "
     "gece modu, hava salinimi (m3/dk), ses seviyesi (dB), zaman ayari (saat), "
     "elektronik kontrol paneli", 4),
    (('eco safe', 'comfort eco', 'seramik isitici', 'fanli isitici', 'oda isitici'),
     "seramik fanli isitma teknolojisi, m2 cinsinden isitma alani, ses seviyesi (dB), "
     "sogutma fonksiyonu, eko-fonksiyon, elektronik zamanlayicili otomatik acilma/kapanma", 4),
]

def _match_category_hints(context_text):
    text = context_text.lower()
    matched = [(hint, m[2] if len(m) > 2 else None) for m in CATEGORY_SPEC_HINTS
               for keywords, hint in [(m[0], m[1])] if any(k in text for k in keywords)]
    hint_text = '; '.join(h for h, _ in matched)
    counts = [c for _, c in matched if c]
    return hint_text, (max(counts) if counts else None)


def extract_tech_bullets(tech_pairs, product_name='', category_context='', count=2):
    """Use Claude Haiku to pick the most catalog-worthy concrete specs from the
    TECHNICAL CHARACTERISTICS table, prioritizing the spec(s) that actually define
    quality for that product type (e.g. steam pressure for an iron, wattage for a
    blender, motor rpm for a razor, airflow for a hair dryer)."""
    if not tech_pairs:
        return []
    try:
        import anthropic
        client = anthropic.Anthropic()
        table_text = '\n'.join(f'{k}: {v}' for k, v in tech_pairs.items())
        priority_hint, priority_count = _match_category_hints(f'{category_context} {product_name}')
        if priority_count:
            count = priority_count
        priority_line = (
            f"Bu urun icin ozellikle su spesifikasyonlara oncelik ver: {priority_hint}.\n"
            if priority_hint else ''
        )
        prompt = (
            f"Urun: {product_name}\n\n"
            f"Asagida bu urunun teknik ozellik tablosu var. Bu urun kategorisi icin "
            f"alıcının en çok önemsediği, kaliteyi belirleyen somut spesifikasyonu/spesifikasyonları sec "
            f"(örnek: ütü için buhar basıncı/buhar çıkışı, blender/mutfak robotu için watt güç, "
            f"traş makinesi için motor hızı veya pil ömrü, saç kurutma makinesi için hava akışı/watt). "
            f"Genel/sıradan özellikleri değil, o kategori için kritik olanı seç. "
            f"{priority_line}"
            f"Tablo ne kadar sınırlı olursa olsun, mutlaka elindeki en somut/sayısal "
            f"{count} veriyi seçip madde haline getir. Asla soru sorma, özür dileme "
            f"veya yorum yazma — sadece {count} kısa madde dön (örnek format: "
            f"'2400W güç', '6 bar buhar basıncı'), satır satır, numara veya tire "
            "koyma, baska hicbir aciklama ekleme.\n\n"
            f"Tablo:\n{table_text}"
        )
        message = client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=150,
            messages=[{"role": "user", "content": prompt}],
        )
        lines = [l.strip(' -•\t') for l in message.content[0].text.strip().split('\n') if l.strip()]
        # Discard hedging/refusal-style responses (long sentences, questions) —
        # a real bullet is short; keep only lines that actually look like one.
        lines = [l for l in lines if len(l) <= 60 and '?' not in l]
        return [(l, '') for l in lines[:count]]
    except Exception as e:
        print(f"Tech bullets error: {e}")
        return []


def trim_name_to_core(name):
    """Cut a Commercial Name down to just the core model name, dropping the
    trailing marketing tagline (e.g. 'X-Trem Cyclonic Effitech®, Torbasız
    Elektrikli Süpürge, Ultra Verimli' -> 'X-Trem Cyclonic Effitech®')."""
    if not name:
        return name
    cut_points = [i for i in (name.find(','), name.find('–'), name.find('—')) if i != -1]
    return (name[:min(cut_points)] if cut_points else name).strip()


def translate_name_to_turkish(name):
    """Swap known English product-type phrases for their Turkish equivalent;
    leave the rest of the name (brand, model, English or not) untouched."""
    if not name:
        return name
    return re.sub(r'vacuum cleaner', 'elektrikli süpürge', name, flags=re.IGNORECASE)

# ─── Endpoints ─────────────────────────────────────────────────────────────────

@app.route('/health', methods=['GET'])
def health():
    load_fonts()
    summary = {cat: len(prods) for cat, prods in CATALOG_STATE.items()}
    return jsonify({'status': 'ok', 'fonts': list(FONT_MAP.keys()), 'state': summary})

@app.route('/finalize', methods=['POST'])
def finalize():
    """Rebuild PDF for a category using provided state JSON."""
    try:
        category    = request.form.get('category')
        state_json  = request.form.get('state_json')

        if not category or not state_json:
            return jsonify({'error': 'category ve state_json gerekli'}), 400

        state = json.loads(state_json)
        products_list = list(state.get(category, {}).values())

        if not products_list:
            return jsonify({'error': f'{category} icin urun yok'}), 400

        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            out_path = tmp.name

        build_pdf(products_list, out_path, category)

        with open(out_path, 'rb') as fh:
            pdf_data = fh.read()
        os.unlink(out_path)

        safe_cat = category.replace(' ', '_').replace('/', '_').replace('&', 'and')
        return jsonify({
            'pdf_base64':    base64.b64encode(pdf_data).decode(),
            'filename':      f'katalog_{safe_cat}.pdf',
            'category':      category,
            'product_count': len(products_list),
        })

    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/add_product', methods=['POST'])
@app.route('/upload', methods=['POST'])
def add_product():
    try:
        load_fonts()
        if 'file' not in request.files:
            return jsonify({'error': 'file eksik'}), 400

        # Restore product state from previous call (Railway restart resilience)
        if 'prev_state_json' in request.form:
            try:
                incoming = json.loads(request.form['prev_state_json'])
                for cat, prods in incoming.items():
                    for ref, prod in prods.items():
                        if ref not in CATALOG_STATE[cat]:
                            CATALOG_STATE[cat][ref] = prod
            except Exception:
                pass

        xlsm_bytes = request.files['file'].read()
        product    = parse_xlsm(xlsm_bytes, filename=request.files['file'].filename)
        category   = product['category'] or 'GENEL'

        CATALOG_STATE[category][product['ref']] = product
        products_list = list(CATALOG_STATE[category].values())

        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            out_path = tmp.name

        build_pdf(products_list, out_path, category)

        with open(out_path, 'rb') as fh:
            pdf_data = fh.read()
        os.unlink(out_path)

        safe_cat = category.replace(' ', '_').replace('/', '_').replace('&', 'and')

        state_snapshot = {cat: dict(prods) for cat, prods in CATALOG_STATE.items()}

        return jsonify({
            'pdf_base64':    base64.b64encode(pdf_data).decode(),
            'filename':      f'katalog_{safe_cat}.pdf',
            'category':      category,
            'product_ref':   product['ref'],
            'product_count': len(products_list),
            'state_json':    state_snapshot,
        })

    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)


@app.route('/find_fonts', methods=['GET'])
def find_fonts():
    import glob
    results = []
    for pattern in [
        '/usr/share/fonts/**/*.ttf',
        '/nix/store/**/dejavu*/*.ttf',
        '/nix/store/**/*.ttf',
        '/home/**/*.ttf',
        '/opt/**/*.ttf',
    ]:
        found = glob.glob(pattern, recursive=True)
        results.extend(found[:5])
    return jsonify({'found': results[:30]})
